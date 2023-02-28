import softest
import logging
import inspect
import csv
from openpyxl import workbook, load_workbook, worksheet


class utils(softest.TestCase):
    def assert_list_item_text(self, list_name, value):
        for s in list_name:
            print("The text in the list is: " + s.text)
            self.soft_assert(self.assertEqual(list_name, value))
            if s.text == value:
                print("the test was passed")
            else:
                print("the test was failed")
        self.assert_all()

    def custom_logger(self, loglevel=logging.DEBUG, file_name=None):
        logger_name = inspect.stack()[1][3]
        logger = logging.getLogger(logger_name)
        logger.setLevel(loglevel)
        fh = logging.FileHandler(file_name)
        formatter = logging.Formatter('%(asctime)s-%(levelname)s-%(name)s:%(message)s', datefmt='%m/%d/%Y %I:%M:%S %p')
        fh.setFormatter(formatter)
        logger.addHandler(fh)
        return logger
    def read_data_from_excelfile(self, file_name, sheet_name):
        datalist = []
        wb = load_workbook(filename=file_name)
        sh = wb[sheet_name]
        row_cout = sh.max_row
        colu_cout = sh.max_colum
        for i in range(2, row_cout+1):
            row = []
            for j in range(1, colu_cout+1):
                row.append(sh.cell(row=i, column=j).value)
            datalist.append(row)
        return datalist
    def read_data_from_csvfile(self, file_name):
        datalist = []
        csv_data = open(file_name, 'r')
        reader = csv.reader(csv_data)
        next(reader)
        for row in reader:
            datalist.append(row)
        return datalist
